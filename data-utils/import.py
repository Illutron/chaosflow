# -*- coding: utf-8 -*-

from __future__ import unicode_literals

import datetime
import os
import glob

from openpyxl.reader.excel import load_workbook
from pymongo import Connection

# see reference at http://api.mongodb.org/python/current/tutorial.html
connection = Connection('127.0.0.1', 27017)
db = connection.chaosflow
stat_buckets = db.stat_buckets

#stat_buckets.ensureIndex({"road": 1}, {unique: true}); # should not work like this, rather check if name is present for the same date.

stat_entries = db.stat_entries

# make it so that if direction_two is empty its a one way road
# get lat lngthrough a reverse geocode lookup

count = 1
path = 'data/'

def add_stat_entry(stat_bucket, hour, amount, direction, type="bike"):
    stat_entry = {
        "stat_bucket": stat_bucket,
        "hour": hour,
        "amount": amount,
        "type": type,
        "direction": direction
        }

    stat_entries.insert(stat_entry)
    

for infile in glob.glob( os.path.join(path, '*.xlsx') ):

    wb = load_workbook(filename=infile)

    # get metadata sheet
    stamopl = wb.get_sheet_by_name(name='Stamopl.')

    # get traffic data sheet
    data = wb.get_sheet_by_name(name='DATA')

    stat_bucket = {}
    for i in range(8):
        # Interpret metadata
        label = stamopl.cell("A{}".format(i+1)).value
        val = stamopl.cell("B{}".format(i+1)).value

        if label == "NUMMER":
            stat_bucket["number"] = val
        elif label == "VEJ":
            stat_bucket["road"] = val
        elif label == "DATO":
            stat_bucket["date"] = val
        elif label == "VEJRET":
            stat_bucket["weather"] = val
        elif label == "VEJRFAKTOR":
            stat_bucket["weather_factor"] = val
        elif label == "ANM:":
            stat_bucket["comments"] = val
        elif label == "RETN. 1":
            stat_bucket["direction_one"] = val
        elif label == "RETN. 2":
            stat_bucket["direction_two"] = val

    current_stat_bucket = stat_buckets.insert(stat_bucket)

    stat_entry = {}

    dir_one = []
    dir_two = []

    # interpret traffic data, for now we only fetch bikes
    if data.cell("H2").value == 62:
        # this is the bike data for direction #1
        for i in range(11):
            dir_one.append(data.cell("H{}".format(3+i)).value)

        if data.cell("H33").value == 62:
            # this is the bike data for direction #2
            for i in range(11): # step through 12 hours of data
                 dir_two.append(data.cell("H{}".format(34+i)).value)

    elif data.cell("M5").value == "_x0001_" or data.cell("M5").value == 62:
        # dir 1 is in m6-m17
        for i in range(11):
             dir_one.append(data.cell("M{}".format(6+i)).value)

        if data.cell("M31").value == "_x0001_" or data.cell("M31").value == 62:
        # dir 2 is in m32-m43
            for i in range(11):
                 dir_two.append(data.cell("M{}".format(32+i)).value)

    if dir_one:
        for i in range(11):
            add_stat_entry(current_stat_bucket, 7+i, dir_one[i], 1)

        if dir_two:
            for i in range(11):
                add_stat_entry(current_stat_bucket, 7+i, dir_two[i], 2)

    count+=1







